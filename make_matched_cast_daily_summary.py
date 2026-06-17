"""
Create daily cast-detail summary for casts present in HM matched output.

Purpose:
    - Read cast details.
    - Read HM matched output.
    - Keep only cast-detail rows whose CAST_NUMBER exists in HM matched output.
    - Remove the time part from OPENING_TIME and group by date.
    - Sum only additive numeric columns.

Default inputs:
    ~/Desktop/HM_Si_Project/input/cast details.csv
    ~/Desktop/HM_Si_Project/input/matched_hm_analysis_output_randomized.xlsx

Default output:
    ~/Desktop/HM_Si_Project/output/Matched_Cast_Daily_Summary.xlsx
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path.home() / "Desktop" / "HM_Si_Project"
DEFAULT_CAST_INPUT = BASE_DIR / "input" / "cast details.csv"
DEFAULT_HM_INPUT = BASE_DIR / "input" / "matched_hm_analysis_output_randomized.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "output" / "Matched_Cast_Daily_Summary.xlsx"
DEFAULT_HM_SHEET = "HM Match Output"
DATE_COLUMN = "OPENING_TIME"

# These columns are totals over the day. Columns like TAPHOLE, CLAY_TYPE,
# Slag_Ratio, CAST_SPEED, pressure, and temperature are not additive.
SUM_COLUMNS = [
    "DURATION",
    "CLAY_QTY",
    "HM_QTY",
]


def normalize_name(value) -> str:
    value = "" if value is None else str(value)
    return re.sub(r"\s+", " ", value.strip())


def normalize_cast_number(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    elif path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
    else:
        raise ValueError(f"Unsupported input file type: {path}")

    df.columns = [normalize_name(c) for c in df.columns]
    return df


def read_hm_matched(path: Path, sheet_name: str | None) -> tuple[pd.DataFrame, str]:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
        df.columns = [normalize_name(c) for c in df.columns]
        return df, "CSV"

    xls = pd.ExcelFile(path)
    sheet = sheet_name or DEFAULT_HM_SHEET
    if sheet not in xls.sheet_names:
        sheet = xls.sheet_names[0]

    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    keywords = {"cast_number", "sentdate", "opening_time", "hm_si", "recipename"}

    best_row = 0
    best_score = -1
    for idx in range(min(10, len(raw))):
        vals = {normalize_name(v).lower() for v in raw.iloc[idx].tolist()}
        score = len(keywords.intersection(vals))
        if score > best_score:
            best_score = score
            best_row = idx

    if best_score >= 2:
        columns = [
            normalize_name(v) if normalize_name(v) else f"Unnamed_{i}"
            for i, v in enumerate(raw.iloc[best_row].tolist())
        ]
        df = raw.iloc[best_row + 1 :].copy()
        df.columns = columns
    else:
        df = pd.read_excel(path, sheet_name=sheet)
        df.columns = [normalize_name(c) for c in df.columns]

    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return df, sheet


def require_columns(df: pd.DataFrame, columns: list[str], label: str) -> None:
    missing = [col for col in columns if col not in df.columns]
    if missing:
        raise ValueError(f"{label} missing required columns: {missing}")


def build_outputs(cast_df: pd.DataFrame, hm_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    require_columns(cast_df, ["CAST_NUMBER", DATE_COLUMN], "Cast details")
    require_columns(hm_df, ["CAST_NUMBER"], "HM matched data")

    cast = cast_df.copy()
    hm = hm_df.copy()
    cast["CAST_NUMBER_KEY"] = cast["CAST_NUMBER"].apply(normalize_cast_number)
    hm["CAST_NUMBER_KEY"] = hm["CAST_NUMBER"].apply(normalize_cast_number)

    matched_cast_numbers = set(hm["CAST_NUMBER_KEY"].dropna())
    matched = cast[cast["CAST_NUMBER_KEY"].isin(matched_cast_numbers)].copy()
    matched = matched.drop(columns=["CAST_NUMBER_KEY"])

    matched["Cast_Date"] = pd.to_datetime(matched[DATE_COLUMN], errors="coerce").dt.date
    matched = matched.dropna(subset=["Cast_Date"]).copy()

    available_sum_cols = [col for col in SUM_COLUMNS if col in matched.columns]
    for col in available_sum_cols:
        matched[col] = pd.to_numeric(matched[col], errors="coerce")

    summary = matched.groupby("Cast_Date", dropna=True).size().reset_index(name="Cast_Count")
    if available_sum_cols:
        sums = (
            matched.groupby("Cast_Date", dropna=True)[available_sum_cols]
            .sum(min_count=1)
            .reset_index()
        )
        sums = sums.rename(columns={col: f"Total_{col}" for col in available_sum_cols})
        summary = summary.merge(sums, on="Cast_Date", how="left")

    summary = summary.sort_values("Cast_Date").reset_index(drop=True)
    matched = matched.sort_values(["Cast_Date", "CAST_NUMBER"]).reset_index(drop=True)
    return matched, summary


def style_output(output: Path) -> None:
    wb = load_workbook(output)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    date_fill = PatternFill("solid", fgColor="DDEBF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                cell.font = Font(name="Arial", size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 11), 28)

    summary = wb["Daily_Cast_Summary"]
    for row in summary.iter_rows(min_row=2, max_row=summary.max_row):
        for cell in row:
            if cell.column == 1:
                cell.fill = date_fill
            else:
                cell.fill = total_fill
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"

    wb.save(output)


def write_output(
    matched: pd.DataFrame,
    summary: pd.DataFrame,
    output: Path,
    cast_input: Path,
    hm_input: Path,
    hm_sheet: str,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                ["Cast Input", str(cast_input)],
                ["HM Matched Input", str(hm_input)],
                ["HM Matched Sheet", hm_sheet],
                ["Match Key", "CAST_NUMBER"],
                ["Date Column", DATE_COLUMN],
                ["Summed Columns", ", ".join([col for col in SUM_COLUMNS if col in matched.columns])],
                ["Ignored Columns", "Non-additive fields like CLAY_TYPE, TAPHOLE, Slag_Ratio, speed, pressure, temperature"],
            ],
            columns=["Field", "Value"],
        ).to_excel(writer, sheet_name="README", index=False)
        matched.to_excel(writer, sheet_name="Matched_Cast_Details", index=False)
        summary.to_excel(writer, sheet_name="Daily_Cast_Summary", index=False)

    style_output(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--cast-input", default=str(DEFAULT_CAST_INPUT), help="Cast details CSV/XLSX file")
    parser.add_argument("--hm-input", default=str(DEFAULT_HM_INPUT), help="HM matched CSV/XLSX file")
    parser.add_argument("--hm-sheet", default=DEFAULT_HM_SHEET, help="HM matched Excel sheet name")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output Excel file")
    args = parser.parse_args()

    cast_input = Path(args.cast_input).expanduser()
    hm_input = Path(args.hm_input).expanduser()
    output = Path(args.output).expanduser()

    cast_df = read_table(cast_input)
    hm_df, hm_sheet = read_hm_matched(hm_input, args.hm_sheet)
    matched, summary = build_outputs(cast_df, hm_df)
    write_output(matched, summary, output, cast_input, hm_input, hm_sheet)

    print("Done.")
    print(f"Matched cast rows: {len(matched)}")
    print(f"Daily summary rows: {len(summary)}")
    print(f"Output saved to: {output.resolve()}")


if __name__ == "__main__":
    main()
