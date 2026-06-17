"""
Create daily average summary from HM matched sheet.

Purpose:
    - Read the HM matched Excel/CSV file.
    - Use OPENING_TIME by default to create a date-only column.
    - Group rows with the same date.
    - Average all useful numeric columns for that date.

Default input:
    ~/Desktop/HM_Si_Project/input/matched_hm_analysis_output_randomized.xlsx

Default output:
    ~/Desktop/HM_Si_Project/output/HM_Daily_Average_Summary.xlsx
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path.home() / "Desktop" / "HM_Si_Project"
DEFAULT_INPUT = BASE_DIR / "input" / "matched_hm_analysis_output_randomized.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "output" / "HM_Daily_Average_Summary.xlsx"
DEFAULT_SHEET = "HM Match Output"
DEFAULT_DATE_COLUMN = "OPENING_TIME"
DEFAULT_START_ROW = 3
DEFAULT_END_ROW = 920

NON_AVERAGE_COLUMNS = {
    "CAST_NUMBER",
    "SentDate",
    "OPENING_TIME",
    "RecipeName",
    "Cast_Date",
}


def normalize_name(value) -> str:
    value = "" if value is None else str(value)
    return re.sub(r"\s+", " ", value.strip())


def read_hm_sheet(
    path: Path,
    sheet_name: str | None,
    start_row: int,
    end_row: int,
) -> tuple[pd.DataFrame, str]:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
        df.columns = [normalize_name(c) for c in df.columns]
        df["__Excel_Row"] = df.index + 2
        df = df[(df["__Excel_Row"] >= start_row) & (df["__Excel_Row"] <= end_row)].copy()
        df = df.drop(columns=["__Excel_Row"])
        return df.dropna(axis=0, how="all").dropna(axis=1, how="all"), "CSV"

    xls = pd.ExcelFile(path)
    sheet = sheet_name or DEFAULT_SHEET
    if sheet not in xls.sheet_names:
        sheet = xls.sheet_names[0]

    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    keywords = {
        "cast_number",
        "sentdate",
        "opening_time",
        "hm_c",
        "hm_si",
        "recipename",
        "injrate",
        "cr",
    }

    best_row = 0
    best_score = -1
    for idx in range(min(10, len(raw))):
        vals = {normalize_name(v).lower() for v in raw.iloc[idx].tolist()}
        score = len(keywords.intersection(vals))
        if score > best_score:
            best_score = score
            best_row = idx

    if best_score >= 3:
        columns = [
            normalize_name(v) if normalize_name(v) else f"Unnamed_{i}"
            for i, v in enumerate(raw.iloc[best_row].tolist())
        ]
        df = raw.iloc[best_row + 1 :].copy()
        df.columns = columns
        excel_rows = df.index + 1
    else:
        df = pd.read_excel(path, sheet_name=sheet)
        df.columns = [normalize_name(c) for c in df.columns]
        excel_rows = df.index + 2

    df["__Excel_Row"] = excel_rows
    df = df[(df["__Excel_Row"] >= start_row) & (df["__Excel_Row"] <= end_row)].copy()
    df = df.drop(columns=["__Excel_Row"])
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return df, sheet


def choose_date_column(df: pd.DataFrame, requested: str) -> str:
    if requested in df.columns:
        return requested
    for candidate in ["OPENING_TIME", "SentDate"]:
        if candidate in df.columns:
            return candidate
    raise ValueError(
        f"Could not find date column. Tried requested={requested!r}, OPENING_TIME, SentDate"
    )


def build_daily_average(df: pd.DataFrame, date_column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    date_col = choose_date_column(df, date_column)
    working = df.copy()
    working["HM_Date"] = pd.to_datetime(working[date_col], errors="coerce").dt.date
    working = working.dropna(subset=["HM_Date"]).copy()

    numeric_cols = []
    for col in working.columns:
        if col in NON_AVERAGE_COLUMNS or col == "HM_Date":
            continue
        converted = pd.to_numeric(working[col], errors="coerce")
        if converted.notna().any():
            working[col] = converted
            numeric_cols.append(col)

    summary = working.groupby("HM_Date", dropna=True).size().reset_index(name="Row_Count")
    if numeric_cols:
        averages = (
            working.groupby("HM_Date", dropna=True)[numeric_cols]
            .mean()
            .reset_index()
            .rename(columns={col: f"Avg_{col}" for col in numeric_cols})
        )
        summary = summary.merge(averages, on="HM_Date", how="left")

    summary = summary.sort_values("HM_Date").reset_index(drop=True)
    return working.sort_values("HM_Date").reset_index(drop=True), summary


def style_output(output: Path) -> None:
    wb = load_workbook(output)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    date_fill = PatternFill("solid", fgColor="DDEBF7")
    count_fill = PatternFill("solid", fgColor="FCE4D6")
    average_fill = PatternFill("solid", fgColor="E2F0D9")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                cell.font = Font(name="Arial", size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 11), 28)

    summary = wb["HM_Daily_Averages"]
    for row in summary.iter_rows(min_row=1, max_row=summary.max_row):
        for cell in row:
            if cell.column == 1:
                cell.fill = date_fill
            elif cell.column == 2:
                cell.fill = count_fill
            else:
                cell.fill = average_fill
            if cell.row == 1:
                cell.font = Font(name="Arial", size=10, bold=True, color="000000")
            elif isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"

    wb.save(output)


def write_output(
    detail: pd.DataFrame,
    summary: pd.DataFrame,
    output: Path,
    input_path: Path,
    source_sheet: str,
    date_column: str,
    start_row: int,
    end_row: int,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                ["HM Input", str(input_path)],
                ["Source Sheet", source_sheet],
                ["Date Column Used", date_column],
                ["Excel Rows Used", f"{start_row} to {end_row}"],
                ["Logic", "Group same date and average numeric HM/recipe values"],
                ["Ignored Columns", ", ".join(sorted(NON_AVERAGE_COLUMNS))],
            ],
            columns=["Field", "Value"],
        ).to_excel(writer, sheet_name="README", index=False)
        detail.to_excel(writer, sheet_name="HM_Filtered_Data", index=False)
        summary.to_excel(writer, sheet_name="HM_Daily_Averages", index=False)

    style_output(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="HM matched Excel/CSV file")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="HM matched Excel sheet name")
    parser.add_argument("--date-column", default=DEFAULT_DATE_COLUMN, help="Date-time column to group by")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output Excel file")
    parser.add_argument("--start-row", type=int, default=DEFAULT_START_ROW, help="First Excel row to include")
    parser.add_argument("--end-row", type=int, default=DEFAULT_END_ROW, help="Last Excel row to include")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser()
    output = Path(args.output).expanduser()

    df, source_sheet = read_hm_sheet(input_path, args.sheet, args.start_row, args.end_row)
    used_date_col = choose_date_column(df, args.date_column)
    detail, summary = build_daily_average(df, used_date_col)
    write_output(
        detail,
        summary,
        output,
        input_path,
        source_sheet,
        used_date_col,
        args.start_row,
        args.end_row,
    )

    print("Done.")
    print(f"HM rows used: {len(detail)}")
    print(f"Daily average rows: {len(summary)}")
    print(f"Date column used: {used_date_col}")
    print(f"Excel rows used: {args.start_row} to {args.end_row}")
    print(f"Output saved to: {output.resolve()}")


if __name__ == "__main__":
    main()
