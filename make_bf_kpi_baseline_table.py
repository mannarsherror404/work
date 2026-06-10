"""
Create Blast Furnace KPI Baseline Table

Purpose:
    Reads your combined/matched Excel file and creates a clean KPI baseline table
    like the old intern screenshot:

    Parameter | UOM | Weighted Average Values | Target Values |
    Optimization Potential | Price Parameter & UoM | Price/Value |
    Savings in INR Cr/year

How to run:
    pip install pandas openpyxl numpy

    python make_bf_kpi_baseline_table.py \
        --input "matched_hm_analysis_output_randomized.xlsx" \
        --sheet "HM Match Output" \
        --output "BF_KPI_Baseline_Table.xlsx"

Notes:
    - Weighted Average Values are calculated from your current matched data.
    - Target Values, Price/Value, and Savings are left blank/editable because
      those need approved plant/costing inputs.
    - Optimization Potential is added as an Excel formula where numeric target
      values are later filled by you.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_COLUMNS = [
    "Parameter",
    "UOM",
    "Weighted Average Values",
    "Target Values",
    "Optimization Potential",
    "Price Parameter & UoM",
    "Price/Value",
    "Savings in INR Cr/year",
]


KPI_CONFIG = [
    # Improvement / quality parameters
    {"section": "Improvement Parameters"},
    {"param": "Hot Metal C", "columns": ["HM_C"], "uom": "%"},
    {"param": "Hot Metal Si", "columns": ["HM_Si"], "uom": "%"},
    {"param": "Hot Metal S", "columns": ["HM_S"], "uom": "%"},
    {"param": "Hot Metal Mn", "columns": ["HM_Mn"], "uom": "%"},
    {"param": "Hot Metal Ti", "columns": ["HM_Ti"], "uom": "%"},
    {"param": "Hot Metal P", "columns": ["HM_P"], "uom": "%"},
    {"param": "Hot Metal Cr", "columns": ["HM_Cr"], "uom": "%"},
    {"param": "Hot Metal Temperature", "columns": ["HM_Temp_Mean_For_Cast", "HMT", "Hot_Metal_Temp"], "uom": "deg C"},

    # Production / recipe / operating parameters
    {"section": "Production Parameters"},
    {"param": "Sinter", "columns": ["Sinter"], "uom": "plant UOM"},
    {"param": "Pellet", "columns": ["Pellet"], "uom": "plant UOM"},
    {"param": "Ore", "columns": ["Ore"], "uom": "plant UOM"},
    {"param": "NC", "columns": ["NC"], "uom": "plant UOM"},
    {"param": "Dolo", "columns": ["Dolo"], "uom": "plant UOM"},
    {"param": "QTZ", "columns": ["QTZ"], "uom": "plant UOM"},
    {"param": "LS", "columns": ["LS"], "uom": "plant UOM"},
    {"param": "SS", "columns": ["SS"], "uom": "plant UOM"},
    {"param": "Oxide Wt", "columns": ["OxideWt"], "uom": "plant UOM"},
    {"param": "CB", "columns": ["CB"], "uom": "plant UOM"},
    {"param": "PCI / Injection Rate", "columns": ["InjRate"], "uom": "kg/THM"},
    {"param": "CR / Coke Rate", "columns": ["CR"], "uom": "kg/THM"},
    {"param": "NCR", "columns": ["NCR"], "uom": "kg/THM"},
    {"param": "Fuel Rate", "columns": ["FuelRate"], "uom": "kg/THM"},
    {"param": "Flux Rate", "columns": ["FluxRate"], "uom": "plant UOM"},
    {"param": "HM per Charge", "columns": ["HMperCh"], "uom": "plant UOM"},
    {"param": "SR", "columns": ["SR"], "uom": "plant UOM"},
    {"param": "B2", "columns": ["B2"], "uom": "ratio"},
    {"param": "Slag Al2O3", "columns": ["SlagAl2O3"], "uom": "%"},
    {"param": "Slag MgO", "columns": ["SlagMgO"], "uom": "%"},
]


def normalize_name(value) -> str:
    value = "" if value is None else str(value)
    return re.sub(r"\s+", " ", value.strip())


def read_matched_sheet(path: Path, sheet_name: str | None) -> tuple[pd.DataFrame, str]:
    xls = pd.ExcelFile(path)
    sheet = sheet_name or "HM Match Output"
    if sheet not in xls.sheet_names:
        sheet = xls.sheet_names[0]

    raw = pd.read_excel(path, sheet_name=sheet, header=None)

    # Detect the real header row. This handles your file where row 1 has
    # colored section headers and row 2 has actual column names.
    best_row = 0
    best_score = -1
    keywords = [
        "CAST_NUMBER",
        "SentDate",
        "OPENING_TIME",
        "HM_C",
        "HM_Si",
        "RecipeName",
        "InjRate",
        "FuelRate",
    ]
    for idx in range(min(8, len(raw))):
        vals = [normalize_name(v) for v in raw.iloc[idx].tolist()]
        score = sum(any(k.lower() == v.lower() for v in vals) for k in keywords)
        if score > best_score:
            best_score = score
            best_row = idx

    if best_score >= 3:
        columns = [
            normalize_name(v) if normalize_name(v) else f"Unnamed_{i}"
            for i, v in enumerate(raw.iloc[best_row].tolist())
        ]
        df = raw.iloc[best_row + 1 :].copy()
        df.columns = columns
    else:
        df = pd.read_excel(path, sheet_name=sheet)
        df.columns = [normalize_name(c) for c in df.columns]

    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return df, sheet


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lower_map = {str(c).lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    for c in candidates:
        for real_col in df.columns:
            if c.lower() in str(real_col).lower():
                return real_col
    return None


def numeric_average(series: pd.Series):
    values = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    if values.empty:
        return None
    return round(float(values.mean()), 4)


def build_kpi_table(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for item in KPI_CONFIG:
        if "section" in item:
            rows.append([item["section"], "", "", "", "", "", "", ""])
            continue

        col = find_column(df, item["columns"])
        if not col:
            continue

        avg = numeric_average(df[col])
        if avg is None:
            continue

        rows.append([
            item["param"],
            item["uom"],
            avg,
            "",
            "",
            "",
            "",
            "",
        ])

    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


def write_formatted_output(kpi: pd.DataFrame, output: Path, source_file: Path, source_sheet: str):
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                ["Source File", str(source_file)],
                ["Source Sheet", source_sheet],
                ["Generated Table", "KPI baseline table only"],
                ["Note", "Fill Target Values and Price/Value after plant/costing approval."],
            ],
            columns=["Field", "Value"],
        ).to_excel(writer, sheet_name="README", index=False)
        kpi.to_excel(writer, sheet_name="KPI_Baseline_Table", index=False)

    wb = load_workbook(output)
    ws = wb["KPI_Baseline_Table"]

    orange = PatternFill("solid", fgColor="C55A11")
    blue = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Arial", size=10)

    for cell in ws[1]:
        cell.fill = orange
        cell.font = Font(name="Arial", size=10, bold=True, color="000000")

    for row_idx in range(2, ws.max_row + 1):
        parameter = ws.cell(row_idx, 1).value
        if parameter in ("Improvement Parameters", "Production Parameters"):
            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row_idx, col_idx)
                c.fill = blue
                c.font = Font(name="Arial", size=10, bold=True)

        # Excel formula for Optimization Potential:
        # Target Values - Weighted Average Values.
        # It stays blank until Target Values is filled.
        target_cell = ws.cell(row_idx, 4).coordinate
        avg_cell = ws.cell(row_idx, 3).coordinate
        opt_cell = ws.cell(row_idx, 5)
        if parameter not in ("Improvement Parameters", "Production Parameters"):
            opt_cell.value = f'=IF(OR({target_cell}="",{avg_cell}=""),"",{target_cell}-{avg_cell})'

    widths = {
        "A": 28,
        "B": 14,
        "C": 20,
        "D": 18,
        "E": 20,
        "F": 24,
        "G": 16,
        "H": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    readme = wb["README"]
    for row in readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = border
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 90

    wb.save(output)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Combined matched Excel file")
    parser.add_argument("--sheet", default="HM Match Output", help="Matched data sheet name")
    parser.add_argument("--output", default="BF_KPI_Baseline_Table.xlsx", help="Output Excel file")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    df, source_sheet = read_matched_sheet(input_path, args.sheet)
    kpi = build_kpi_table(df)
    write_formatted_output(kpi, output_path, input_path, source_sheet)
    print(f"Created: {output_path.resolve()}")


if __name__ == "__main__":
    main()
