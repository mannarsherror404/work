"""
Create PCI and Coke Rate condition analysis from matched HM data.

Purpose:
    - Create one table sorted by high PCI / injection rate with recipe details.
    - Create one table sorted by low coke rate with recipe details.
    - Summarize average recipe and operating conditions by PCI and coke-rate class.

Default input:
    ~/Desktop/HM_Si_Project/input/matched_hm_analysis_output_randomized.xlsx

Default output:
    ~/Desktop/HM_Si_Project/output/PCI_CR_Condition_Analysis.xlsx

Default data range:
    Excel rows 3 to 920 only. This avoids footer/total rows.
"""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path.home() / "Desktop" / "HM_Si_Project"
DEFAULT_INPUT = BASE_DIR / "input" / "matched_hm_analysis_output_randomized.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "output" / "PCI_CR_Condition_Analysis.xlsx"
DEFAULT_SHEET = "HM Match Output"
DEFAULT_START_ROW = 3
DEFAULT_END_ROW = 920

PCI_CANDIDATES = ["InjRate", "PCI", "PCI Rate", "Injection Rate", "PCI / Injection Rate"]
CR_CANDIDATES = ["CR", "Coke Rate", "CR / Coke Rate"]

CONDITION_PRIORITY = [
    "RecipeName",
    "S",
    "P",
    "O",
    "Sinter",
    "Pellet",
    "Ore",
    "NC",
    "Dolo",
    "QTZ",
    "LS",
    "SS",
    "OxideWt",
    "CB",
    "NCR",
    "FuelRate",
    "FluxRate",
    "HMperCh",
    "SR",
    "B2",
    "SlagAl2O3",
    "SlagMgO",
]


def normalize_name(value) -> str:
    value = "" if value is None else str(value)
    return re.sub(r"\s+", " ", value.strip())


def compact_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def read_matched_sheet(
    path: Path,
    sheet_name: str | None,
    start_row: int,
    end_row: int,
) -> tuple[pd.DataFrame, str]:
    xls = pd.ExcelFile(path)
    sheet = sheet_name or DEFAULT_SHEET
    if sheet not in xls.sheet_names:
        sheet = xls.sheet_names[0]

    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    keywords = {
        "cast_number",
        "sentdate",
        "opening_time",
        "hm_si",
        "recipename",
        "injrate",
        "cr",
    }
